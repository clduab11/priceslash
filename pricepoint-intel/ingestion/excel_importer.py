"""
Excel data importer for PricePoint Intel.
Handles .xlsx and .xls files for bulk vendor pricing data.
"""

import logging
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional, Any, Callable
from dataclasses import dataclass, field
import uuid

# Optional openpyxl import - gracefully handle if not installed
try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from .validator import DataValidator, ValidationResult
from .csv_importer import ImportResult, COLUMN_MAPPINGS

logger = logging.getLogger(__name__)


class ExcelImporter:
    """
    Imports Excel data into PricePoint Intel.
    Supports .xlsx files with multi-sheet handling.
    """

    def __init__(
        self,
        validator: Optional[DataValidator] = None,
    ):
        if not OPENPYXL_AVAILABLE:
            logger.warning(
                "openpyxl not installed. Excel import will not be available. "
                "Install with: pip install openpyxl"
            )
        self.validator = validator or DataValidator()

    def import_sku_products(
        self,
        file_path: str,
        sheet_name: Optional[str] = None,
        column_mapping: Optional[Dict[str, str]] = None,
        on_row_imported: Optional[Callable[[Dict], None]] = None,
    ) -> ImportResult:
        """Import SKU products from Excel."""
        return self._import_file(
            file_path=file_path,
            data_type="sku_product",
            validate_fn=self.validator.validate_sku_product,
            sheet_name=sheet_name,
            column_mapping=column_mapping,
            on_row_imported=on_row_imported,
        )

    def import_vendors(
        self,
        file_path: str,
        sheet_name: Optional[str] = None,
        column_mapping: Optional[Dict[str, str]] = None,
        on_row_imported: Optional[Callable[[Dict], None]] = None,
    ) -> ImportResult:
        """Import vendors from Excel."""
        return self._import_file(
            file_path=file_path,
            data_type="vendor",
            validate_fn=self.validator.validate_vendor,
            sheet_name=sheet_name,
            column_mapping=column_mapping,
            on_row_imported=on_row_imported,
        )

    def import_vendor_pricing(
        self,
        file_path: str,
        sheet_name: Optional[str] = None,
        column_mapping: Optional[Dict[str, str]] = None,
        on_row_imported: Optional[Callable[[Dict], None]] = None,
    ) -> ImportResult:
        """Import vendor pricing from Excel."""
        return self._import_file(
            file_path=file_path,
            data_type="vendor_pricing",
            validate_fn=self.validator.validate_vendor_pricing,
            sheet_name=sheet_name,
            column_mapping=column_mapping,
            on_row_imported=on_row_imported,
        )

    def import_geographic_markets(
        self,
        file_path: str,
        sheet_name: Optional[str] = None,
        column_mapping: Optional[Dict[str, str]] = None,
        on_row_imported: Optional[Callable[[Dict], None]] = None,
    ) -> ImportResult:
        """Import geographic markets from Excel."""
        return self._import_file(
            file_path=file_path,
            data_type="geographic_market",
            validate_fn=self.validator.validate_geographic_market,
            sheet_name=sheet_name,
            column_mapping=column_mapping,
            on_row_imported=on_row_imported,
        )

    def import_distribution_centers(
        self,
        file_path: str,
        sheet_name: Optional[str] = None,
        column_mapping: Optional[Dict[str, str]] = None,
        on_row_imported: Optional[Callable[[Dict], None]] = None,
    ) -> ImportResult:
        """Import distribution centers from Excel."""
        return self._import_file(
            file_path=file_path,
            data_type="distribution_center",
            validate_fn=self.validator.validate_distribution_center,
            sheet_name=sheet_name,
            column_mapping=column_mapping,
            on_row_imported=on_row_imported,
        )

    def _import_file(
        self,
        file_path: str,
        data_type: str,
        validate_fn: Callable,
        sheet_name: Optional[str] = None,
        column_mapping: Optional[Dict[str, str]] = None,
        on_row_imported: Optional[Callable[[Dict], None]] = None,
    ) -> ImportResult:
        """
        Generic Excel file import method.

        Args:
            file_path: Path to the Excel file
            data_type: Type of data being imported
            validate_fn: Validation function to use
            sheet_name: Optional sheet name (uses first sheet if not specified)
            column_mapping: Optional custom column mapping
            on_row_imported: Optional callback for each imported row

        Returns:
            ImportResult with details of the import
        """
        result = ImportResult(
            success=True,
            source_file=file_path,
            started_at=datetime.now(),
        )

        if not OPENPYXL_AVAILABLE:
            result.success = False
            result.add_error(
                0,
                "openpyxl is not installed. Install with: pip install openpyxl"
            )
            result.completed_at = datetime.now()
            return result

        path = Path(file_path)
        if not path.exists():
            result.success = False
            result.add_error(0, f"File not found: {file_path}")
            result.completed_at = datetime.now()
            return result

        try:
            # Load workbook
            workbook = openpyxl.load_workbook(
                path,
                read_only=True,
                data_only=True,  # Get computed values, not formulas
            )

            # Select sheet
            if sheet_name:
                if sheet_name not in workbook.sheetnames:
                    result.success = False
                    result.add_error(
                        0,
                        f"Sheet '{sheet_name}' not found. "
                        f"Available sheets: {workbook.sheetnames}"
                    )
                    result.completed_at = datetime.now()
                    return result
                sheet = workbook[sheet_name]
            else:
                sheet = workbook.active

            logger.info(f"Importing {data_type} from {file_path} (sheet: {sheet.title})")

            # Get headers from first row
            headers = []
            for cell in sheet[1]:
                headers.append(str(cell.value) if cell.value else "")

            if not any(headers):
                result.success = False
                result.add_error(0, "Excel file appears to be empty or has no headers")
                result.completed_at = datetime.now()
                workbook.close()
                return result

            # Build column mapping
            mapping = self._build_column_mapping(
                headers, data_type, column_mapping
            )

            logger.debug(f"Column mapping: {mapping}")

            # Process rows (skip header)
            for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                # Skip completely empty rows
                if not any(cell is not None for cell in row):
                    result.records_skipped += 1
                    continue

                result.records_total += 1

                try:
                    # Create row dict
                    row_dict = {headers[i]: cell for i, cell in enumerate(row) if i < len(headers)}

                    # Map columns
                    mapped_data = self._map_row(row_dict, mapping)

                    # Generate ID if not present
                    id_field = self._get_id_field(data_type)
                    if id_field and not mapped_data.get(id_field):
                        mapped_data[id_field] = str(uuid.uuid4())

                    # Validate
                    validation = validate_fn(mapped_data, row_index=row_num)
                    result.records_processed += 1

                    if validation.is_valid:
                        final_data = validation.cleaned_data or mapped_data
                        result.imported_data.append(final_data)
                        result.records_success += 1

                        if on_row_imported:
                            on_row_imported(final_data)
                    else:
                        result.records_failed += 1
                        for error in validation.errors:
                            result.add_error(row_num, error.message, mapped_data)

                    # Add warnings
                    for warning in validation.warnings:
                        result.add_warning(row_num, warning.message)

                except Exception as e:
                    result.records_failed += 1
                    result.add_error(row_num, f"Error processing row: {str(e)}")
                    logger.exception(f"Error processing row {row_num}")

            workbook.close()

        except Exception as e:
            result.success = False
            result.add_error(0, f"Error reading Excel file: {str(e)}")
            logger.exception("Error during Excel import")

        result.completed_at = datetime.now()

        if result.records_failed > 0:
            result.success = False

        logger.info(
            f"Import complete: {result.records_success}/{result.records_total} records imported"
        )

        return result

    def _build_column_mapping(
        self,
        headers: List[str],
        data_type: str,
        custom_mapping: Optional[Dict[str, str]] = None,
    ) -> Dict[str, str]:
        """Build a mapping from Excel columns to data fields."""
        if custom_mapping:
            return custom_mapping

        mapping = {}
        field_mappings = COLUMN_MAPPINGS.get(data_type, {})

        # Normalize headers for matching
        normalized_headers = {
            str(h).lower().strip().replace(" ", "_"): h
            for h in headers if h
        }

        for field, aliases in field_mappings.items():
            for alias in aliases:
                norm_alias = alias.lower().strip().replace(" ", "_")
                if norm_alias in normalized_headers:
                    mapping[field] = normalized_headers[norm_alias]
                    break

        return mapping

    def _map_row(self, row: Dict[str, Any], mapping: Dict[str, str]) -> Dict[str, Any]:
        """Map an Excel row to data fields using the column mapping."""
        mapped = {}
        for field, column in mapping.items():
            if column in row:
                value = row[column]
                # Handle None and empty values
                if value is None or (isinstance(value, str) and value.strip() == ""):
                    value = None
                # Handle datetime objects
                elif hasattr(value, 'isoformat'):
                    value = value.isoformat()
                mapped[field] = value
        return mapped

    def _get_id_field(self, data_type: str) -> Optional[str]:
        """Get the ID field name for a data type."""
        id_fields = {
            "sku_product": "sku_id",
            "vendor": "vendor_id",
            "vendor_pricing": "pricing_id",
            "geographic_market": "market_id",
            "distribution_center": "center_id",
        }
        return id_fields.get(data_type)

    def get_sheet_names(self, file_path: str) -> List[str]:
        """Get list of sheet names in an Excel file."""
        if not OPENPYXL_AVAILABLE:
            return []

        path = Path(file_path)
        if not path.exists():
            return []

        try:
            workbook = openpyxl.load_workbook(path, read_only=True)
            names = workbook.sheetnames
            workbook.close()
            return names
        except Exception as e:
            logger.error(f"Error reading sheet names: {e}")
            return []

    def preview(
        self,
        file_path: str,
        data_type: str,
        sheet_name: Optional[str] = None,
        max_rows: int = 5,
        column_mapping: Optional[Dict[str, str]] = None,
    ) -> Dict[str, Any]:
        """
        Preview an Excel file without importing.

        Args:
            file_path: Path to the Excel file
            data_type: Type of data to validate as
            sheet_name: Optional sheet name
            max_rows: Maximum rows to preview
            column_mapping: Optional custom column mapping

        Returns:
            Dict with preview information
        """
        if not OPENPYXL_AVAILABLE:
            return {"error": "openpyxl not installed"}

        path = Path(file_path)
        if not path.exists():
            return {"error": f"File not found: {file_path}"}

        try:
            workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)

            if sheet_name:
                if sheet_name not in workbook.sheetnames:
                    return {
                        "error": f"Sheet '{sheet_name}' not found",
                        "available_sheets": workbook.sheetnames,
                    }
                sheet = workbook[sheet_name]
            else:
                sheet = workbook.active

            # Get headers
            headers = [
                str(cell.value) if cell.value else ""
                for cell in sheet[1]
            ]

            mapping = self._build_column_mapping(
                headers, data_type, column_mapping
            )

            # Get sample rows
            sample_rows = []
            for row_num, row in enumerate(
                sheet.iter_rows(min_row=2, max_row=max_rows + 1, values_only=True)
            ):
                row_dict = {headers[i]: cell for i, cell in enumerate(row) if i < len(headers)}
                mapped = self._map_row(row_dict, mapping)
                sample_rows.append(mapped)

            # Count total rows
            total_rows = sheet.max_row - 1 if sheet.max_row else 0

            workbook.close()

            return {
                "file_path": str(path),
                "sheet_name": sheet.title,
                "available_sheets": workbook.sheetnames if hasattr(workbook, 'sheetnames') else [],
                "headers": headers,
                "column_mapping": mapping,
                "unmapped_columns": [h for h in headers if h and h not in mapping.values()],
                "unmapped_fields": [
                    f for f in COLUMN_MAPPINGS.get(data_type, {}).keys()
                    if f not in mapping
                ],
                "sample_rows": sample_rows,
                "total_rows": total_rows,
            }

        except Exception as e:
            return {"error": str(e)}


def import_workbook(
    file_path: str,
    validator: Optional[DataValidator] = None,
    sheet_mapping: Optional[Dict[str, str]] = None,
) -> Dict[str, ImportResult]:
    """
    Import multiple sheets from an Excel workbook.

    Args:
        file_path: Path to the Excel file
        validator: Optional DataValidator instance
        sheet_mapping: Dict mapping sheet names to data types
            Example: {"Products": "sku_product", "Pricing": "vendor_pricing"}

    Returns:
        Dict mapping sheet names to ImportResult
    """
    importer = ExcelImporter(validator=validator)
    results = {}

    if not sheet_mapping:
        # Try to auto-detect based on sheet names
        sheet_names = importer.get_sheet_names(file_path)
        sheet_mapping = {}

        for name in sheet_names:
            lower_name = name.lower()
            if any(k in lower_name for k in ["sku", "product"]):
                sheet_mapping[name] = "sku_product"
            elif any(k in lower_name for k in ["vendor", "supplier"]):
                sheet_mapping[name] = "vendor"
            elif any(k in lower_name for k in ["price", "pricing", "cost"]):
                sheet_mapping[name] = "vendor_pricing"
            elif any(k in lower_name for k in ["market", "region", "geo"]):
                sheet_mapping[name] = "geographic_market"
            elif any(k in lower_name for k in ["distribution", "center", "dc", "warehouse"]):
                sheet_mapping[name] = "distribution_center"

    for sheet_name, data_type in sheet_mapping.items():
        if data_type == "sku_product":
            results[sheet_name] = importer.import_sku_products(file_path, sheet_name)
        elif data_type == "vendor":
            results[sheet_name] = importer.import_vendors(file_path, sheet_name)
        elif data_type == "vendor_pricing":
            results[sheet_name] = importer.import_vendor_pricing(file_path, sheet_name)
        elif data_type == "geographic_market":
            results[sheet_name] = importer.import_geographic_markets(file_path, sheet_name)
        elif data_type == "distribution_center":
            results[sheet_name] = importer.import_distribution_centers(file_path, sheet_name)

    return results
